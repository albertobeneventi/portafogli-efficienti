"""
exporter.py — Export Excel, CSV e PDF per portafogli.
Supporta:
  - PDF professionale con 3 portafogli (Max Sharpe, Min Vol, Black-Litterman)
  - Excel standard con pesi e metriche
  - CSV AdvisorElite  (ISIN, Amount%)
  - Excel AdvisorElite (virtual-positions-template: ISIN, Amount, Quantity, Price, Currency, Portfolio)
"""

import io
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, Image, HRFlowable, KeepTogether, PageBreak,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os

    _FONT_NAME = "Helvetica"
    _FONT_BOLD = "Helvetica-Bold"
    _unicode_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
    ]
    _FONT_BOLD_PATHS = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "C:/Windows/Fonts/calibrib.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    for _fp in _unicode_paths:
        if _os.path.exists(_fp):
            try:
                pdfmetrics.registerFont(TTFont("UniFont", _fp))
                _FONT_NAME = "UniFont"
                break
            except Exception:
                pass
    if _FONT_NAME != "Helvetica":
        for _fp in _FONT_BOLD_PATHS:
            if _os.path.exists(_fp):
                try:
                    pdfmetrics.registerFont(TTFont("UniFont-Bold", _fp))
                    _FONT_BOLD = "UniFont-Bold"
                    break
                except Exception:
                    pass
        if _FONT_BOLD == "Helvetica-Bold":
            _FONT_BOLD = _FONT_NAME   # fallback: stesso font normale

    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    _FONT_NAME = "Helvetica"
    _FONT_BOLD = "Helvetica-Bold"

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Colori
NAVY      = "#1A2C54"
NAVY_RL   = colors.HexColor(NAVY)   if HAS_REPORTLAB else None
GOLD_RL   = colors.HexColor("#C9A84C") if HAS_REPORTLAB else None
GRAY_LIGHT = colors.HexColor("#F5F7FA") if HAS_REPORTLAB else None
GRAY_MID   = colors.HexColor("#E2E8F0") if HAS_REPORTLAB else None
GREEN_RL   = colors.HexColor("#2E7D32") if HAS_REPORTLAB else None


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def plotly_to_png(fig, width: int = 1100, height: int = 500) -> Optional[bytes]:
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        pass
    try:
        import plotly.io as pio
        return pio.to_image(fig, format="png", width=width, height=height, scale=2)
    except Exception:
        return None


def _safe_str(text) -> str:
    """Converte in stringa sicura per Helvetica (ASCII/Latin-1) o lascia invariata per UniFont."""
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    if _FONT_NAME != "Helvetica":
        return text
    _MAP = {
        "★": "*", "☆": "o", "€": "EUR",
        "–": "-", "—": "-",
        "‘": "'", "’": "'",
        "“": '"', "”": '"',
        "\xe0": "a", "\xe8": "e", "\xe9": "e",
        "\xec": "i", "\xf2": "o", "\xf9": "u",
    }
    for src, dst in _MAP.items():
        text = text.replace(src, dst)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _stars(n) -> str:
    if n is None:
        return "-"
    try:
        n = int(float(str(n)))
        if not (1 <= n <= 5):
            return "-"
        s = "*" if _FONT_NAME == "Helvetica" else "★"
        e = "o" if _FONT_NAME == "Helvetica" else "☆"
        return s * n + e * (5 - n)
    except Exception:
        return "-"


def _fmt_pct(v, decimals: int = 2) -> str:
    try:
        return f"{float(v):.{decimals}f}%"
    except Exception:
        return "-"


def _fmt_num(v, decimals: int = 3) -> str:
    try:
        return f"{float(v):.{decimals}f}"
    except Exception:
        return "-"


# ---------------------------------------------------------------------------
# EXCEL EXPORT (standard)
# ---------------------------------------------------------------------------

def export_portfolio_excel(
    weights: dict,
    metrics: dict,
    fund_df: Optional[pd.DataFrame] = None,
    corr_df: Optional[pd.DataFrame] = None,
    price_dict: Optional[dict] = None,
    title: str = "Portafoglio",
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        weights_df = pd.DataFrame([
            {"ISIN": isin, "Peso (%)": round(w * 100, 2)}
            for isin, w in weights.items() if w > 0
        ])
        if fund_df is not None and not fund_df.empty:
            try:
                _cols = [c for c in ["isin", "nome", "classificazione", "casa",
                                     "perf_1y", "perf_3y", "volatilita",
                                     "rating_fida", "retrocessione"] if c in fund_df.columns]
                fund_info = fund_df[_cols].copy()
                fund_info.columns = [c.replace("_", " ").title() for c in _cols]
                fund_info = fund_info.rename(columns={"Isin": "ISIN"})
                weights_df = weights_df.merge(fund_info, on="ISIN", how="left")
            except Exception:
                pass
        weights_df.to_excel(writer, sheet_name="Pesi", index=False)
        metrics_df = pd.DataFrame([{"Metrica": k, "Valore": v} for k, v in metrics.items()])
        metrics_df.to_excel(writer, sheet_name="Metriche", index=False)
        if corr_df is not None and not corr_df.empty:
            corr_df.to_excel(writer, sheet_name="Correlazioni")
        if price_dict:
            try:
                pd.DataFrame(price_dict).to_excel(writer, sheet_name="Serie Storiche")
            except Exception:
                pass
        # Stile
        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment
        navy_fill  = PatternFill(start_color="1A2C54", end_color="1A2C54", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        for sn in writer.sheets:
            ws = writer.sheets[sn]
            for cell in ws[1]:
                cell.fill = navy_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ADVISORELITE CSV
# ---------------------------------------------------------------------------

def export_advisorelite_csv(weights: dict) -> bytes:
    """
    Formato AdvisorElite CSV: ISIN, Amount (% interi, somma = 100).
    """
    rows = [(isin, round(w * 100, 4))
            for isin, w in weights.items() if w > 0.0001]
    # Ribilancia per garantire somma = 100
    total = sum(r[1] for r in rows)
    if total > 0:
        rows = [(isin, round(pct / total * 100, 4)) for isin, pct in rows]
    lines = ["ISIN,Amount,"]
    for isin, pct in sorted(rows, key=lambda x: -x[1]):
        lines.append(f"{isin},{pct}")
    return "\n".join(lines).encode("utf-8")


# ---------------------------------------------------------------------------
# ADVISORELITE EXCEL (virtual-positions-template)
# ---------------------------------------------------------------------------

def export_advisorelite_excel(
    weights: dict,
    total_amount: float = 100_000.0,
    currency: str = "EUR",
    portfolio_name: str = "Portafoglio",
    fund_pool: Optional[dict] = None,
) -> bytes:
    """
    Formato virtual-positions-template AdvisorElite:
    Colonne: ISIN, Amount, Quantity, Price, Currency, Portfolio
    Amount  = importo in EUR (peso% × total_amount)
    Quantity = Amount / Price (se disponibile, altrimenti = Amount)
    Price   = 1.0 se non disponibile (placeholder)
    """
    if not HAS_OPENPYXL:
        return b""

    rows = []
    _pool = fund_pool or {}
    for isin, w in sorted(weights.items(), key=lambda x: -x[1]):
        if w <= 0.0001:
            continue
        amount = round(w * total_amount, 2)
        price  = 1.0
        info   = _pool.get(isin, {})
        # Tenta di recuperare prezzo/nav
        for price_key in ["nav", "price", "prezzo", "last_price"]:
            p = info.get(price_key)
            if p and float(p) > 0:
                price = float(p)
                break
        quantity = round(amount / price, 4) if price > 0 else amount
        rows.append({
            "ISIN":      isin,
            "Amount":    amount,
            "Quantity":  quantity,
            "Price":     price,
            "Currency":  currency,
            "Portfolio": portfolio_name,
        })

    buf = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    # Header stile AdvisorElite (righe descrittive + row ISIN)
    headers = ["ISIN", "Amount", "Quantity", "Price", "Currency", "Portfolio"]
    desc_row = [
        "Free entry input (Mandatory)",
        "Numeric input (Mandatory)\n",
        "Numeric input (Mandatory)\n",
        "Numeric input (Mandatory)\n",
        "Free entry input (Mandatory)\nFund or ETF currency must be introduced",
        "Not completable input",
    ]
    navy_fill  = PatternFill(start_color="1A2C54", end_color="1A2C54", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    data_font  = Font(name="Calibri", size=10)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Riga 1: descrizioni
    for col_idx, desc in enumerate(desc_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=desc)
        cell.alignment = center
        cell.font = Font(name="Calibri", size=9, color="666666", italic=True)

    # Riga 2: header
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=hdr)
        cell.fill  = navy_fill
        cell.font  = white_font
        cell.alignment = center

    # Righe dati
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.font      = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
            cell.border    = border
        # Zebra
        if (row_idx - 3) % 2 == 1:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = \
                    PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    # Larghezze colonne
    widths = [20, 14, 14, 10, 12, 22]
    for col_idx, w_val in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w_val

    # Foglio Data (valute)
    ws_data = wb.create_sheet("Data")
    ws_data["A1"] = "Currencies"
    for r, cur in enumerate(["EUR", "USD", "GBP", "CHF"], 2):
        ws_data.cell(row=r, column=1, value=cur)

    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF EXPORT — 3 PORTAFOGLI
# ---------------------------------------------------------------------------

def export_portfolio_pdf(
    weights: dict,
    metrics: dict,
    title: str = "Report Portafoglio",
    chart_bytes: Optional[bytes] = None,
    frontier_fig=None,
    fund_df: Optional[pd.DataFrame] = None,
    fund_pool: Optional[dict] = None,
    include_stars: bool = True,
    # Portafogli aggiuntivi
    weights_minvol: Optional[dict] = None,
    metrics_minvol: Optional[dict] = None,
    weights_bl: Optional[dict] = None,
    metrics_bl: Optional[dict] = None,
    bl_views: Optional[dict] = None,
    primary_title: str = "Portafoglio Max Sharpe",
    cone_bytes: Optional[bytes] = None,
    cone_portfolios: Optional[list] = None,   # [{"label":str,"mu":float,"sigma":float}]
    cone_capitale: float = 100_000,
    cone_orizzonte: int = 10,
    cone_reliability_pct: Optional[int] = None,
) -> bytes:
    """
    Genera PDF con:
    - Grafico frontiera efficiente
    - Sezione Max Sharpe
    - Sezione Min Volatilità (se fornita)
    - Sezione Black-Litterman (se fornita)
    - Footer disclaimer
    """
    if not HAS_REPORTLAB:
        return b""

    if chart_bytes is None and frontier_fig is not None:
        chart_bytes = plotly_to_png(frontier_fig)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.8 * cm, leftMargin=1.8 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    # Larghezza utile ≈ 29.7 - 3.6 = 26.1 cm
    _PAGE_W = 26.1 * cm
    styles = getSampleStyleSheet()

    def _p(name, **kw):
        # Se fontName è in kw (es. per h1/h2 bold), usa quello; altrimenti default
        fn = kw.pop("fontName", _FONT_NAME)
        return ParagraphStyle(name, parent=styles["Normal"], fontName=fn, **kw)

    # NB: ParagraphStyle eredita leading=12 da styles["Normal"] se non specificato
    # esplicitamente. Con fontSize > 12 questo causa sovrapposizione tra le righe
    # (es. titolo a 20pt con leading 12pt si sovrappone al paragrafo successivo).
    # Ogni stile qui sotto imposta leading ≈ 1.25x fontSize.
    title_style   = _p("T",  fontSize=20, leading=25, textColor=NAVY_RL, spaceAfter=4)
    sub_style     = _p("S",  fontSize=9,  leading=12, textColor=colors.HexColor("#666666"), spaceAfter=10)
    h1_style      = _p("H1", fontSize=14, leading=18, textColor=NAVY_RL, spaceBefore=16, spaceAfter=6,
                        fontName=_FONT_BOLD)
    h2_style      = _p("H2", fontSize=11, leading=14, textColor=NAVY_RL, spaceBefore=10, spaceAfter=4,
                        fontName=_FONT_BOLD)
    small         = _p("Sm", fontSize=8,  leading=10, textColor=colors.HexColor("#555555"), spaceAfter=3)
    note_style    = _p("N",  fontSize=8,  leading=10, textColor=colors.HexColor("#888888"),
                        leftIndent=10, spaceAfter=6)
    cell_style    = _p("Cell", fontSize=8, leading=9.5, wordWrap="CJK")
    cell_h_style  = _p("CellH", fontSize=8.5, leading=10, textColor=colors.white,
                        fontName=_FONT_BOLD, wordWrap="CJK")

    story = []

    # ── Intestazione ──────────────────────────────────────────────────────
    story.append(Paragraph(_safe_str(title), title_style))
    story.append(Paragraph(
        _safe_str(f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}  |  "
                  "Portafogli Efficienti  |  Solo uso interno"),
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY_RL, spaceAfter=10))

    # ── Grafico ───────────────────────────────────────────────────────────
    if chart_bytes:
        story.append(Paragraph("Frontiera Efficiente", h1_style))
        try:
            img = Image(io.BytesIO(chart_bytes), width=24.0 * cm, height=10 * cm)
            story.append(img)
            story.append(Paragraph(
                _safe_str("Curva blu = frontiera efficiente  |  "
                          "Stella rossa = Max Sharpe  |  Diamante = Min Varianza"),
                note_style,
            ))
        except Exception as e:
            story.append(Paragraph(f"Grafico non disponibile: {e}", small))
        story.append(Spacer(1, 0.4 * cm))

    # ── Costruisce info_map per lookup nome/classificazione ───────────────
    _info_map: dict = {}
    if fund_pool:
        _info_map.update(fund_pool)
    if fund_df is not None and not fund_df.empty and "isin" in fund_df.columns:
        for _, r in fund_df.iterrows():
            isin = str(r.get("isin", ""))
            if isin and isin not in _info_map:
                _info_map[isin] = r.to_dict()

    def _portfolio_section(
        h_title: str,
        h_color,
        w_dict: dict,
        m_dict: dict,
        note: str = "",
        new_page: bool = True,
    ):
        """Aggiunge una sezione portafoglio (titolo + metriche + tabella pesi).
        Ogni sezione parte da una pagina nuova per evitare titoli a metà pagina
        e sovrapposizioni con la sezione precedente."""
        if new_page:
            story.append(PageBreak())

        _header_block = [
            Paragraph(_safe_str(h_title), ParagraphStyle(
                h_title, parent=styles["Normal"],
                fontName=_FONT_BOLD, fontSize=13, textColor=h_color,
                spaceBefore=0, spaceAfter=5,
            )),
        ]
        if note:
            _header_block.append(Paragraph(_safe_str(note), note_style))

        # Metriche
        m_data = [["Metrica", "Valore"]]
        for k, v in m_dict.items():
            m_data.append([_safe_str(str(k)), _safe_str(str(v))])
        mt = Table(m_data, colWidths=[12 * cm, 9 * cm])
        mt.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), NAVY_RL),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, -1), _FONT_NAME),
            ("FONTNAME",      (0, 0), (-1, 0),  _FONT_BOLD),
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRAY_LIGHT]),
            ("GRID",          (0, 0), (-1, -1), 0.4, GRAY_MID),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        _header_block.append(mt)
        # Titolo + nota + tabella metriche restano insieme: se non c'entrano
        # nella pagina corrente, ReportLab li sposta in blocco alla pagina
        # successiva invece di spezzarli (niente titoli "orfani" a fondo pagina).
        story.append(KeepTogether(_header_block))
        story.append(Spacer(1, 0.3 * cm))

        # Tabella pesi
        story.append(Paragraph("Composizione", h2_style))
        has_stars = include_stars and any(
            _info_map.get(isin, {}).get("rating_fida") is not None
            for isin in w_dict if w_dict.get(isin, 0) > 0
        )
        if has_stars:
            hdr   = ["ISIN", "Nome / Fondo", "Asset Class", "Peso %", "★ FIDA", "Perf 3Y %"]
            col_w = [2.8 * cm, 10.5 * cm, 6.0 * cm, 2.0 * cm, 2.2 * cm, 2.2 * cm]
        else:
            hdr   = ["ISIN", "Nome / Fondo", "Asset Class", "Peso %", "Perf 3Y %"]
            col_w = [2.8 * cm, 11.9 * cm, 7.0 * cm, 2.0 * cm, 2.4 * cm]

        # Header e celle "Nome/Fondo"/"Asset Class" come Paragraph: il testo va
        # a capo dentro la colonna invece di sforare nella colonna successiva
        # (causa delle sovrapposizioni quando il nome del fondo è lungo).
        w_data = [[Paragraph(_safe_str(h), cell_h_style) for h in hdr]]
        for isin, w in sorted(w_dict.items(), key=lambda x: -x[1]):
            if w <= 0:
                continue
            info  = _info_map.get(isin, {})
            nome  = _safe_str(str(info.get("nome", info.get("FONDO AZIMUT", isin))))
            cl    = _safe_str(str(info.get("classificazione", info.get("categoria", "-"))))
            perf  = info.get("perf_3y")
            p_str = f"{float(perf):.1f}%" if perf is not None else "-"
            nome_p = Paragraph(nome, cell_style)
            cl_p   = Paragraph(cl, cell_style)
            if has_stars:
                w_data.append([isin, nome_p, cl_p, f"{w * 100:.1f}%",
                                _stars(info.get("rating_fida")), p_str])
            else:
                w_data.append([isin, nome_p, cl_p, f"{w * 100:.1f}%", p_str])

        _ts = [
            ("BACKGROUND",    (0, 0), (-1, 0), NAVY_RL),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, -1), _FONT_NAME),
            ("FONTNAME",      (0, 0), (-1, 0),  _FONT_BOLD),
            ("FONTSIZE",      (0, 0), (-1, -1), 8.5),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRAY_LIGHT]),
            ("GRID",          (0, 0), (-1, -1), 0.35, GRAY_MID),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN",         (3, 0), (-1, -1), "CENTER"),
        ]
        wt = Table(w_data, colWidths=col_w, repeatRows=1)
        wt.setStyle(TableStyle(_ts))
        story.append(wt)
        if has_stars:
            legend = ("Rating FIDA: ***** = 5 stelle" if _FONT_NAME == "Helvetica"
                      else "Rating FIDA: ★★★★★ = 5 stelle (top)")
            story.append(Paragraph(legend, small))
        story.append(Spacer(1, 0.5 * cm))

    # ── 1. Sezione principale (Max Sharpe nel report Frontiera Efficiente,
    #      oppure il portafoglio passato per export singoli es. Qualità) ─────
    _primary_note = (
        "Massimizza il rapporto rendimento/rischio (Indice di Sharpe)."
        if primary_title == "Portafoglio Max Sharpe" else ""
    )
    _portfolio_section(
        h_title=primary_title,
        h_color=NAVY_RL,
        w_dict=weights,
        m_dict=metrics,
        note=_primary_note,
        # La prima sezione resta sempre sulla stessa pagina dell'intestazione
        # (ed eventuale grafico): mai una pagina 1 con solo titolo, anche se
        # il grafico non viene generato (es. kaleido non disponibile).
        new_page=False,
    )

    # ── 2. Min Volatilità ─────────────────────────────────────────────────
    if weights_minvol and metrics_minvol:
        _portfolio_section(
            h_title="Portafoglio Min Volatilità",
            h_color=GREEN_RL,
            w_dict=weights_minvol,
            m_dict=metrics_minvol,
            note="Minimizza la volatilità del portafoglio (approccio difensivo).",
        )

    # ── 3. Black-Litterman ────────────────────────────────────────────────
    if weights_bl and metrics_bl:
        note_bl = "Incorpora le aspettative di mercato via Score Qualità (Black-Litterman)."
        if bl_views:
            views_str = "  |  ".join(
                f"{isin}: {v:+.1f}%" for isin, v in list(bl_views.items())[:6]
            )
            note_bl += f"\nView principali: {views_str}"
        _portfolio_section(
            h_title="Portafoglio Black-Litterman",
            h_color=GOLD_RL,
            w_dict=weights_bl,
            m_dict=metrics_bl,
            note=note_bl,
        )

    # ── Cono di Ibbotson ─────────────────────────────────────────────────
    if cone_bytes:
        story.append(PageBreak())
        _rel_str = (f"  |  Attendibilita' della stima: {cone_reliability_pct}%"
                    if cone_reliability_pct is not None else "")
        story.append(Paragraph(
            _safe_str(f"Cono di Ibbotson — Proiezione futura{_rel_str}"), h1_style))
        story.append(Paragraph(
            _safe_str(
                "Il grafico mostra il possibile valore futuro del portafoglio nel tempo. "
                "La linea centrale e' il caso piu' probabile (mediana). "
                "Le bande mostrano la dispersione attorno ad esso: "
                "quella interna copre il 68% degli scenari storici, "
                "quella esterna il 95%."
            ),
            note_style,
        ))
        try:
            cone_img = Image(io.BytesIO(cone_bytes), width=24.0 * cm, height=9 * cm)
            story.append(cone_img)
        except Exception as _ce:
            story.append(Paragraph(f"Grafico cono non disponibile: {_ce}", small))

        # ── Tabella scenari ───────────────────────────────────────────────
        if cone_portfolios:
            import math as _math
            _years = [y for y in (1, 3, 5, 10) if y <= cone_orizzonte]
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph("Scenari per orizzonte temporale", h2_style))
            story.append(Paragraph(
                _safe_str(
                    "Valori in euro del portafoglio nei diversi scenari. "
                    "'Scenario molto sfavorevole': solo il 2.5% dei casi reali ha fatto peggio. "
                    "'Caso centrale': meta' dei casi finisce sopra, meta' sotto. "
                    "'Scenario molto favorevole': solo il 2.5% dei casi supera questo valore."
                ),
                note_style,
            ))
            for _cp in cone_portfolios:
                _mu, _sig = _cp["mu"], _cp["sigma"]
                _mu_log = _mu - 0.5 * _sig ** 2
                _lbl = _cp.get("label", "")
                if len(cone_portfolios) > 1:
                    story.append(Paragraph(_safe_str(_lbl), small))

                _hdr = ["Anni",
                        "Scenario molto\nsfavorevole",
                        "Scenario\nsfavorevole",
                        "Caso centrale\n(più probabile)",
                        "Scenario\nfavorevole",
                        "Scenario molto\nfavorevole"]
                _tbl_data = [[Paragraph(_safe_str(h), cell_h_style) for h in _hdr]]
                for _t in _years:
                    _row_vals = [
                        cone_capitale * _math.exp((_mu_log - 2 * _sig) * _t),
                        cone_capitale * _math.exp((_mu_log - _sig) * _t),
                        cone_capitale * _math.exp(_mu_log * _t),
                        cone_capitale * _math.exp((_mu_log + _sig) * _t),
                        cone_capitale * _math.exp((_mu_log + 2 * _sig) * _t),
                    ]
                    _tbl_data.append(
                        [Paragraph(_safe_str(str(_t)), cell_style)]
                        + [Paragraph(_safe_str(f"€{v:,.0f}"), cell_style)
                           for v in _row_vals]
                    )

                _col_w = [1.5 * cm] + [(_PAGE_W - 1.5 * cm) / 5] * 5
                _tbl = Table(_tbl_data, colWidths=_col_w, repeatRows=1)
                _tbl.setStyle(TableStyle([
                    ("BACKGROUND",  (0, 0), (-1, 0), NAVY_RL),
                    ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#F0F4FF")]),
                    ("GRID",        (0, 0), (-1, -1), 0.4, GRAY_MID),
                    ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
                    ("ALIGN",       (0, 0), (0, -1),  "CENTER"),
                    ("FONTSIZE",    (0, 0), (-1, -1),  8),
                    # Colonna mediana in grassetto
                    ("FONTNAME",    (3, 1), (3, -1), _FONT_BOLD),
                ]))
                story.append(_tbl)
                story.append(Spacer(1, 0.25 * cm))

    # ── Footer ────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY_MID))
    story.append(Paragraph(
        _safe_str(
            "Documento generato da Portafogli Efficienti  |  Solo uso interno.  "
            "Non costituisce consulenza finanziaria. Le performance passate non sono "
            "garanzia di risultati futuri."
        ),
        small,
    ))

    doc.build(story)
    return buf.getvalue()
